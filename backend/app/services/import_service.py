"""Promovidos importer: parse messy multi-sheet XLSX into Registro-ready dicts."""
from __future__ import annotations

import hashlib
import os
import re
from typing import Optional

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core import crypto
from app.models.registro import Registro
from app.services.audit_service import record_audit

_GENERIC_SHEETS = {"C1", "A", "HOJA1", "HOJA 1", "SHEET1"}

# Registro column widths (String(n)). Values are capped to these so a stray
# oversized cell (or a column-misaligned row) can never abort a batch commit
# with a Postgres StringDataRightTruncation.
_MAXLEN = {
    "nombre_completo": 255, "direccion": 500, "colonia": 255,
    "telefono": 40, "estructura": 120, "promotor": 160, "observacion": 1000,
}
_SECCION_RE = re.compile(r"\d{1,6}")
# Clave de elector INE: exactly 18 alphanumeric characters.
_CLAVE_RE = re.compile(r"[A-Za-z0-9]{18}")


def _clean(v) -> str:
    return re.sub(r"\s+", " ", str(v).strip()) if v is not None else ""


def _norm_clave(v) -> Optional[str]:
    """Normalize a clave de elector cell: strip all whitespace, upper-case, and
    accept only exactly-18 alphanumeric chars. Anything else → None (absent or
    garbage), so a bad cell never stores an invalid clave."""
    s = re.sub(r"\s+", "", _clean(v)).upper()
    return s if _CLAVE_RE.fullmatch(s) else None


def _find_clave_col(ws, hdr: int) -> Optional[int]:
    """0-based column index of a 'CLAVE DE ELECTOR' header, or None.

    Detected by NAME (not fixed position) because the clave column is optional
    and appears in varying positions across promotor templates. Scans the two
    header rows (label may span both) and returns the first column whose joined
    label contains 'CLAVE'. Index aligns with the ``values_only`` row tuples.
    """
    labels: dict[int, str] = {}
    for row in ws.iter_rows(min_row=hdr, max_row=hdr + 1):
        for idx, cell in enumerate(row):
            txt = _clean(cell.value).upper()
            if txt:
                labels[idx] = f"{labels.get(idx, '')} {txt}".strip()
    for idx in sorted(labels):
        if "CLAVE" in labels[idx]:
            return idx
    return None


def _fit(v: Optional[str], key: str) -> Optional[str]:
    n = _MAXLEN.get(key)
    if v is not None and n is not None and len(v) > n:
        return v[:n]
    return v


def _norm_seccion(v) -> Optional[str]:
    """A sección is a short number (e.g. 4138). Excel may hand us an int/float.
    Returns the normalized digits, or None if absent/garbage (a non-numeric
    value here means the row's columns are misaligned)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return str(int(v))
    s = _clean(v)
    return s or None


def _edad_from(dia, mes, anio, ref_year: int = 2026) -> Optional[int]:
    try:
        y = int(float(anio))
    except (TypeError, ValueError):
        return None
    if y < 100:  # 2-digit year
        y = 1900 + y if y > 25 else 2000 + y
    if not (1900 <= y <= ref_year):
        return None
    return ref_year - y


def _find_header_row(ws) -> Optional[int]:
    # Use the 1-based row index (min_row=1) rather than ``row[0].row`` — in
    # read-only mode a blank leading cell is an EmptyCell with no ``.row``.
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8), start=1):
        joined = " ".join(_clean(c.value).upper() for c in row)
        if "PRIMER APELLIDO" in joined and "NOMBRE" in joined:
            return i
    return None


def _file_label(path: str) -> str:
    base = os.path.splitext(os.path.basename(path))[0]
    return re.sub(r"_Mayus$", "", base, flags=re.IGNORECASE).strip()


def parse_workbook(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    estructura = _file_label(path)
    out: list[dict] = []
    for ws in wb.worksheets:
        hdr = _find_header_row(ws)
        if hdr is None:
            continue
        promotor = _clean(ws.title)
        if promotor.upper() in _GENERIC_SHEETS:
            promotor = estructura
        # columns are fixed by the standard template (see spec §5):
        # 2 ap1, 3 ap2, 4 nombre, 5 dia, 6 mes, 7 anio, 8 calle, 9 num,
        # 10 colonia, 11 seccion, 12 telefono
        # clave de elector is OPTIONAL and position-variable → detected by name.
        clave_idx = _find_clave_col(ws, hdr)
        for row in ws.iter_rows(min_row=hdr + 2, values_only=True):
            row = list(row) + [None] * (12 - len(row))
            ap1, ap2, nombre = _clean(row[1]), _clean(row[2]), _clean(row[3])
            if not (ap1 or ap2 or nombre):
                continue  # empty / spacer row
            seccion = _norm_seccion(row[10])
            # A present-but-non-numeric sección means this file's columns are
            # misaligned for this row (e.g. a colonia name landed in the sección
            # cell). Skip it rather than importing garbage / crashing the batch.
            if seccion is not None and not _SECCION_RE.fullmatch(seccion):
                continue
            clave = (_norm_clave(row[clave_idx])
                     if clave_idx is not None and clave_idx < len(row) else None)
            nombre_completo = _clean(f"{nombre} {ap1} {ap2}")
            calle, num = _clean(row[7]), _clean(row[8])
            direccion = _clean(f"{calle} {num}") or None
            tel = re.sub(r"\D", "", _clean(row[11])) or None
            dia, mes, anio = row[4], row[5], row[6]
            edad = _edad_from(dia, mes, anio)
            nac = "/".join(_clean(x) for x in (dia, mes, anio) if _clean(x))
            observacion = f"nac: {nac}" if nac else None
            out.append({
                "nombre_completo": _fit(nombre_completo, "nombre_completo"),
                "direccion": _fit(direccion, "direccion"),
                "colonia": _fit(_clean(row[9]) or None, "colonia"),
                "seccion": seccion,
                "telefono": _fit(tel, "telefono"),
                "edad": edad,
                "observacion": _fit(observacion, "observacion"),
                "promotor": _fit(promotor, "promotor"),
                "estructura": _fit(estructura, "estructura"),
                "clave": clave,
                "_sheet": ws.title,
                "_row": None,  # row index attached by caller for client_uuid
            })
    wb.close()
    return out


def _client_uuid(path: str, sheet: str, idx: int) -> str:
    key = f"{os.path.basename(path)}|{sheet}|{idx}"
    return hashlib.sha1(key.encode("utf-8")).hexdigest()[:32]


def import_rows(db: Session, *, organization_id: str, campaign_id: str, path: str,
                actor_id: Optional[str] = None) -> dict:
    """Idempotently import promovidos from ``path`` into ``Registro``.

    Never logs or prints PII — only counts. Writes one ``registro.import``
    AuditLog row per call (batch-level, not per-record).
    """
    rows = parse_workbook(path)
    leidas = len(rows)
    importadas = 0
    duplicadas = 0
    actualizadas = 0  # existing rows backfilled with a clave they were missing
    # stable per-sheet counter for deterministic client_uuid
    per_sheet: dict[str, int] = {}
    for r in rows:
        sheet = r["_sheet"]
        idx = per_sheet.get(sheet, 0)
        per_sheet[sheet] = idx + 1
        cuid = _client_uuid(path, sheet, idx)
        clave = r.get("clave")
        existing = db.execute(
            select(Registro).where(
                Registro.campaign_id == campaign_id,
                Registro.client_uuid == cuid,
            )
        ).scalar_one_or_none()
        if existing is not None:
            # Backfill a clave onto an already-imported row that lacks one; never
            # overwrite an existing clave. Everything else is an idempotent dupe.
            if clave and not existing.clave_elector_enc:
                existing.clave_elector_enc = crypto.encrypt_clave(clave)
                existing.clave_masked = crypto.mask_clave(clave)
                actualizadas += 1
            else:
                duplicadas += 1
            continue
        db.add(Registro(
            organization_id=organization_id,
            campaign_id=campaign_id,
            activista_id=None,
            nombre_completo=r["nombre_completo"],
            seccion=r["seccion"],
            direccion=r["direccion"],
            colonia=r["colonia"],
            telefono=r["telefono"],
            edad=r["edad"],
            estructura=r["estructura"],
            promotor=r["promotor"],
            observacion=r["observacion"],
            clave_elector_enc=crypto.encrypt_clave(clave) if clave else None,
            clave_masked=crypto.mask_clave(clave) if clave else None,
            consentimiento=True,
            aviso_version="import-papel-2024",
            client_uuid=cuid,
        ))
        importadas += 1
    file_ref = hashlib.sha1(os.path.basename(path).encode("utf-8")).hexdigest()[:16]
    record_audit(
        db, action="registro.import", actor_id=actor_id,
        organization_id=organization_id, entity_type="registro_batch",
        entity_id=file_ref,
        meta={"leidas": leidas, "importadas": importadas,
              "duplicadas": duplicadas, "actualizadas": actualizadas},
    )
    db.commit()
    return {"leidas": leidas, "importadas": importadas,
            "duplicadas": duplicadas, "actualizadas": actualizadas}
