"""Parser de Excel de promovidos + import idempotente."""
import openpyxl
from sqlalchemy import func, select

from app.models.audit_log import AuditLog
from app.models.registro import Registro
from app.services import import_service
from tests.conftest import ALPHA_CAMPAIGN_ID, TestingSessionLocal


def _make_xlsx(path, header_row=1):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ALAN URIEL RAMIREZ"
    r = header_row
    ws.cell(r, 1, "N.P."); ws.cell(r, 2, "PRIMER APELLIDO"); ws.cell(r, 3, "SEGUNDO APELLIDO")
    ws.cell(r, 4, "NOMBRE"); ws.cell(r, 5, "FECHA DE NACIMIENTO"); ws.cell(r, 8, "DOMICILIO")
    ws.cell(r, 12, "TELÉFONO CON WHATSAPP")
    ws.cell(r+1, 5, "DIA"); ws.cell(r+1, 6, "MES"); ws.cell(r+1, 7, "AÑO")
    ws.cell(r+1, 8, "CALLE"); ws.cell(r+1, 9, "#"); ws.cell(r+1, 10, "BARRIO/COLONIA")
    ws.cell(r+1, 11, "SECCIÓN")
    # data rows
    ws.cell(r+2, 1, 1); ws.cell(r+2, 2, "LEÓN"); ws.cell(r+2, 3, "ALCARAZ"); ws.cell(r+2, 4, "PEDRO")
    ws.cell(r+2, 5, 2); ws.cell(r+2, 6, 3); ws.cell(r+2, 7, 1988)
    ws.cell(r+2, 8, "C. MADERO"); ws.cell(r+2, 9, 506); ws.cell(r+2, 10, "BO. SAN FRANCISCO")
    ws.cell(r+2, 11, 4132); ws.cell(r+2, 12, "7226127261")
    # 2-digit year row
    ws.cell(r+3, 1, 2); ws.cell(r+3, 2, "GONZALEZ"); ws.cell(r+3, 3, "DAVILA"); ws.cell(r+3, 4, "ALBERTO")
    ws.cell(r+3, 5, 3); ws.cell(r+3, 6, 6); ws.cell(r+3, 7, 71)
    ws.cell(r+3, 11, 4130); ws.cell(r+3, 12, "7223478883")
    # empty row (only N.P.)
    ws.cell(r+4, 1, 3)
    wb.save(path)


def test_parse_skips_misaligned_rows_and_caps_lengths(tmp_path):
    """A row whose 'sección' cell holds non-numeric text (column misalignment,
    as in FORMATO DE LLENADO FERNANDO GAMA) is skipped, not imported/crashed.
    Oversized string cells are capped to their column width."""
    p = tmp_path / "FORMATO DE LLENADO RARO_Mayus.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PROMOTOR X"
    ws.cell(1, 2, "PRIMER APELLIDO"); ws.cell(1, 4, "NOMBRE"); ws.cell(1, 11, "SECCIÓN")
    ws.cell(2, 11, "SECCIÓN")
    # valid row: numeric sección + an oversized colonia (>255 chars)
    ws.cell(3, 2, "PEREZ"); ws.cell(3, 3, "LOPEZ"); ws.cell(3, 4, "ANA")
    ws.cell(3, 10, "X" * 400); ws.cell(3, 11, 4121)
    # misaligned row: a colonia name landed in the sección cell → must be skipped
    ws.cell(4, 2, "GONZALEZ"); ws.cell(4, 4, "LUIS"); ws.cell(4, 11, "LA CONCEPCION")
    wb.save(str(p))

    rows = import_service.parse_workbook(str(p))
    assert len(rows) == 1                       # misaligned row dropped
    assert rows[0]["seccion"] == "4121"
    assert len(rows[0]["colonia"]) == 255       # capped to column width


def test_parse_maps_columns_and_edad(tmp_path):
    p = tmp_path / "ACTIVISMO CULTURA_Mayus.xlsx"
    _make_xlsx(str(p), header_row=1)
    rows = import_service.parse_workbook(str(p))
    assert len(rows) == 2  # empty row skipped
    r0 = rows[0]
    assert r0["nombre_completo"] == "PEDRO LEÓN ALCARAZ"
    assert r0["seccion"] == "4132"
    assert r0["telefono"] == "7226127261"
    assert r0["edad"] == 2026 - 1988
    assert r0["promotor"] == "ALAN URIEL RAMIREZ"
    assert r0["estructura"] == "ACTIVISMO CULTURA"
    assert r0["observacion"].startswith("nac: ")
    assert rows[1]["edad"] == 2026 - 1971  # 2-digit year 71 → 1971


def test_parse_header_on_row_3(tmp_path):
    p = tmp_path / "EMANUEL_Mayus.xlsx"
    _make_xlsx(str(p), header_row=3)
    rows = import_service.parse_workbook(str(p))
    assert len(rows) == 2 and rows[0]["seccion"] == "4132"


def test_import_rows_idempotent_and_audited(tmp_path):
    p = tmp_path / "ACTIVISMO CULTURA_Mayus.xlsx"
    _make_xlsx(str(p), header_row=1)
    db = TestingSessionLocal()
    try:
        org_id = db.execute(select(Registro.organization_id).limit(1)).scalar()  # may be None
        # use the Alpha org id from a seeded user instead:
        from app.models.user import User
        org_id = db.execute(select(User.organization_id).where(
            User.email == "coord@alpha.gov")).scalar_one()

        res1 = import_service.import_rows(db, organization_id=org_id,
                                          campaign_id=ALPHA_CAMPAIGN_ID, path=str(p))
        assert res1["importadas"] == 2
        n1 = db.execute(select(func.count()).select_from(Registro).where(
            Registro.promotor == "ALAN URIEL RAMIREZ")).scalar_one()
        assert n1 == 2

        # re-run → no duplicates
        res2 = import_service.import_rows(db, organization_id=org_id,
                                          campaign_id=ALPHA_CAMPAIGN_ID, path=str(p))
        assert res2["importadas"] == 0 and res2["duplicadas"] == 2
        n2 = db.execute(select(func.count()).select_from(Registro).where(
            Registro.promotor == "ALAN URIEL RAMIREZ")).scalar_one()
        assert n2 == 2

        # one batch-audit row per import call
        n_audit = db.execute(select(func.count()).select_from(AuditLog).where(
            AuditLog.action == "registro.import")).scalar_one()
        assert n_audit >= 1
    finally:
        db.close()


def test_import_rows_audit_entity_id_is_not_pii(tmp_path):
    """AuditLog.entity_id must be a non-PII, <=36-char hash — never the (PII,
    length-unbounded) filename. Real filenames in docs/data/separados/ are
    person names and can exceed the String(36) column, which crashes on
    Postgres (StringDataRightTruncation) even though SQLite silently accepts
    it. Counts belong in ``meta``, not the identifier."""
    import hashlib

    basename = "DAVID CESAR CORZA MONTES DE OCA LARGO NOMBRE_Mayus.xlsx"
    assert len(basename) > 36
    p = tmp_path / basename
    _make_xlsx(str(p), header_row=1)
    db = TestingSessionLocal()
    try:
        from app.models.user import User
        org_id = db.execute(select(User.organization_id).where(
            User.email == "coord@alpha.gov")).scalar_one()

        res = import_service.import_rows(db, organization_id=org_id,
                                          campaign_id=ALPHA_CAMPAIGN_ID, path=str(p))
        assert res["importadas"] == 2

        # Fetch THIS import's audit row deterministically by its expected
        # entity_id hash — AuditLog.id is a UUID (not a sequential key), so
        # order_by(id).first() is NOT "most recent" and picks a wrong row in the
        # full suite where other tests also write registro.import audits.
        expected_ref = hashlib.sha1(basename.encode("utf-8")).hexdigest()[:16]
        row = db.execute(
            select(AuditLog).where(
                AuditLog.action == "registro.import",
                AuditLog.entity_id == expected_ref,
            )
        ).scalars().one()
        assert row is not None
        assert len(row.entity_id) <= 36
        assert row.entity_id != basename
        assert row.meta is not None
        assert row.meta["importadas"] == 2
        assert row.meta["leidas"] == 2
        assert row.meta["duplicadas"] == 0
    finally:
        db.close()


def _make_xlsx_clave(path, sheet="PROMOTOR CLAVE", con_clave=True):
    """Workbook with an optional 'CLAVE DE ELECTOR' column at col 13."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.cell(1, 2, "PRIMER APELLIDO"); ws.cell(1, 3, "SEGUNDO APELLIDO")
    ws.cell(1, 4, "NOMBRE"); ws.cell(1, 11, "SECCIÓN")
    ws.cell(1, 13, "CLAVE DE ELECTOR")
    ws.cell(2, 11, "SECCIÓN")
    ws.cell(3, 2, "PEREZ"); ws.cell(3, 3, "LOPEZ"); ws.cell(3, 4, "ANA"); ws.cell(3, 11, 4121)
    if con_clave:
        ws.cell(3, 13, "PRLPAN80010112M400")  # 18 alphanumeric
    wb.save(path)


def test_parse_reads_clave_de_elector_column(tmp_path):
    p = tmp_path / "CON CLAVE_Mayus.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PROMOTOR X"
    ws.cell(1, 2, "PRIMER APELLIDO"); ws.cell(1, 4, "NOMBRE")
    ws.cell(1, 11, "SECCIÓN"); ws.cell(1, 13, "CLAVE DE ELECTOR")
    ws.cell(2, 11, "SECCIÓN")
    ws.cell(3, 2, "PEREZ"); ws.cell(3, 4, "ANA"); ws.cell(3, 11, 4121)
    ws.cell(3, 13, "PRLPAN80010112M400")           # valid 18-char clave
    ws.cell(4, 2, "GOMEZ"); ws.cell(4, 4, "LUIS"); ws.cell(4, 11, 4122)
    ws.cell(4, 13, "ABC123")                        # too short → None
    wb.save(str(p))
    rows = import_service.parse_workbook(str(p))
    assert len(rows) == 2
    assert rows[0]["clave"] == "PRLPAN80010112M400"
    assert rows[1]["clave"] is None


def test_parse_without_clave_column_yields_none(tmp_path):
    """Backward compat: templates without a clave column set clave=None."""
    p = tmp_path / "SIN CLAVE_Mayus.xlsx"
    _make_xlsx(str(p), header_row=1)
    rows = import_service.parse_workbook(str(p))
    assert all(r["clave"] is None for r in rows)


def test_import_encrypts_clave_and_backfills_existing(tmp_path):
    from app.core import crypto
    from app.models.user import User
    db = TestingSessionLocal()
    try:
        org_id = db.execute(select(User.organization_id).where(
            User.email == "coord@alpha.gov")).scalar_one()

        # 1) First import WITHOUT clave → row created, no ciphertext.
        p_no = tmp_path / "PROMO SIN_Mayus.xlsx"
        _make_xlsx_clave(str(p_no), con_clave=False)
        res0 = import_service.import_rows(db, organization_id=org_id,
                                          campaign_id=ALPHA_CAMPAIGN_ID, path=str(p_no))
        assert res0["importadas"] == 1
        reg = db.execute(select(Registro).where(
            Registro.campaign_id == ALPHA_CAMPAIGN_ID,
            Registro.nombre_completo == "ANA PEREZ LOPEZ")).scalar_one()
        assert reg.clave_elector_enc is None and reg.clave_masked is None

        # 2) Re-import SAME file (same basename/sheet/row → same client_uuid) WITH
        #    clave → existing row is backfilled (not duplicated), decrypts back.
        p_yes = tmp_path / "PROMO SIN_Mayus.xlsx"  # same name → same client_uuid
        _make_xlsx_clave(str(p_yes), con_clave=True)
        res1 = import_service.import_rows(db, organization_id=org_id,
                                          campaign_id=ALPHA_CAMPAIGN_ID, path=str(p_yes))
        assert res1["actualizadas"] == 1 and res1["importadas"] == 0
        db.refresh(reg)
        assert reg.clave_elector_enc is not None
        assert crypto.decrypt_clave(bytes(reg.clave_elector_enc)) == "PRLPAN80010112M400"

        # 3) Re-import again → already has clave → duplicada, not re-updated.
        res2 = import_service.import_rows(db, organization_id=org_id,
                                          campaign_id=ALPHA_CAMPAIGN_ID, path=str(p_yes))
        assert res2["actualizadas"] == 0 and res2["duplicadas"] == 1
    finally:
        db.close()
